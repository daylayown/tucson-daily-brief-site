#!/usr/bin/env python3
"""
school_spending.py — Arizona Auditor General school-district spending, Pima County.

Downloads the AG's statewide "School District Spending" data file (one XLSX,
all 238 districts) and extracts the 14 Pima County districts into a structured
JSON archive. No AI calls anywhere: every figure is read from a named column
and every derived number is computed here. ("Derive, don't ask.")

Source: A.R.S. § 41-1279.03 requires the Auditor General to monitor what share
of every school dollar reaches the classroom. Published annually.
  landing: https://www.azauditor.gov/arizona-school-district-spending-analysis-and-data-file-fiscal-year-2025
  file:    /sites/default/files/2026-02/AZ_School_District_Spending_FY25_Data_File.xlsx

⚠️ THE METRIC TRAP — read before writing any copy off this data.
The workbook carries TWO different "classroom" percentages and they differ by
~16 points:
  * col 17 "Instruction"  = 0.521 statewide  <- THE headline number. This is the
    Auditor General's Instructional Spending Percentage (ISP), the "52.1 cents
    of every dollar" figure every other outlet quotes. Col 27 (Current ISP)
    matches it exactly.
  * col 16 "Classroom spending" = 0.680 statewide <- a BROADER definition
    (instruction + student support + instruction support).
Reporting col 16 as "classroom dollars" would be wrong by 16 points against the
number readers will compare it to. This module exposes both, named
unambiguously, and `isp` is the one that means "classroom dollars".

FY2025 context: statewide ISP 52.1% is the LOWEST since monitoring began in
2001 (peak 58.6% in 2004). Verified in-file, not from news coverage.

Usage:
    python3 school_spending.py                 # fetch + write data/school_spending.json
    python3 school_spending.py --summary       # print the Pima table
    python3 school_spending.py --no-fetch      # reuse the cached workbook
"""
import argparse
import json
import sys
from datetime import date
from pathlib import Path

import requests

SITE_DIR = Path(__file__).resolve().parent
DATA_DIR = SITE_DIR / "data"
CACHE = DATA_DIR / "az-school-spending-fy25.xlsx"
OUT_JSON = DATA_DIR / "school_spending.json"

LANDING = ("https://www.azauditor.gov/"
           "arizona-school-district-spending-analysis-and-data-file-fiscal-year-2025")
XLSX_URL = ("https://www.azauditor.gov/sites/default/files/2026-02/"
            "AZ_School_District_Spending_FY25_Data_File.xlsx")
SHEET = "FY25 Data"
FISCAL_YEAR = 2025
UA = ("Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) "
      "Chrome/120 Safari/537.36")

# Column indexes (0-based) into the "FY25 Data" sheet. Row 1 = section headers,
# row 2 = column headers, row 3 = State of Arizona, rows 4+ = districts.
COLS = {
    "district": 0,
    "county": 2,
    "location": 7,
    "schools": 8,
    "students": 9,
    "size": 10,
    "enrollment_change_5yr": 11,
    "special_ed_pct": 12,
    "english_learner_pct": 13,
    # spending mix (FY2025)
    "classroom_broad_pct": 16,     # instruction + student support + instr support
    "isp": 17,                     # ***the "classroom dollars" headline metric***
    "student_support_pct": 19,
    "instruction_support_pct": 20,
    "nonclassroom_pct": 21,
    "administration_pct": 22,
    "plant_operations_pct": 23,
    "food_service_pct": 24,
    "transportation_pct": 25,
    # ISP history
    "isp_prior_year": 26,
    "isp_current": 27,
    "isp_highest_fy": 28,
    "isp_highest": 29,
    "isp_lowest_fy": 30,
    "isp_lowest": 31,
    # per-student spending
    "per_student_total_fy24": 141,
    "per_student_total": 142,
    "per_student_peer_avg": 144,
    "per_student_state_avg": 146,
    "admin_per_student": 150,
    # teachers
    "teacher_salary": 184,
    "teacher_salary_vs_state": 186,
    "teacher_experience_yrs": 190,
    "teachers_first_3yrs_pct": 191,
    "salary_first_3yrs": 192,
    "salary_4th_plus": 194,
    "students_per_teacher": 195,
}

# Districts TDB covers as a beat, keyed to SCHOOL-DATA-FEASIBILITY.md's crosswalk.
# Anything else in Pima is still archived, just not flagged as a covered beat.
COVERED = {
    "Tucson Unified School District": "TUSD",
    "Sunnyside Unified School District": "Sunnyside",
    "Vail Unified School District": "Vail",
    "Marana Unified School District": "Marana",
    "Amphitheater Unified School District": "Amphitheater",
    "Sahuarita Unified School District": "Sahuarita",
    "Flowing Wells Unified School District": "Flowing Wells",
    "Catalina Foothills Unified School District": "Catalina Foothills",
    "Tanque Verde Unified School District": "Tanque Verde",
}


def fetch(force=True):
    DATA_DIR.mkdir(exist_ok=True)
    if CACHE.exists() and not force:
        return CACHE
    r = requests.get(XLSX_URL, headers={"User-Agent": UA, "Referer": LANDING},
                     timeout=120)
    r.raise_for_status()
    ct = r.headers.get("content-type", "")
    if "spreadsheet" not in ct and "excel" not in ct:
        sys.exit(f"ERROR: unexpected content-type {ct!r} — AG may have moved the file. "
                 f"Re-check {LANDING}")
    if len(r.content) < 50_000:
        sys.exit(f"ERROR: workbook only {len(r.content)}b — refusing to overwrite cache")
    CACHE.write_bytes(r.content)
    return CACHE


def _num(v):
    return v if isinstance(v, (int, float)) else None


def _coerce(v):
    """Some cells arrive as numeric strings ('16.3'); normalise so downstream
    consumers get real numbers. Genuine 'N/A' text is left alone."""
    if isinstance(v, str):
        s = v.strip().replace(",", "").replace("$", "")
        if s and s.lstrip("-").replace(".", "", 1).isdigit():
            f = float(s)
            return int(f) if f.is_integer() else f
    return v


def extract(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"ERROR: sheet {SHEET!r} missing; found {wb.sheetnames}")
    ws = wb[SHEET]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    header = rows[1]

    # Guard: the column map is positional, so verify two anchors before trusting it.
    if "Instruction" != str(header[COLS["isp"]]).strip():
        sys.exit(f"ERROR: col {COLS['isp']} is {header[COLS['isp']]!r}, expected "
                 "'Instruction'. AG changed the layout — re-map COLS before use.")
    if "District name" != str(header[COLS["district"]]).strip():
        sys.exit("ERROR: col 0 is not 'District name' — layout changed.")

    def pack(row):
        d = {k: _coerce(row[i]) for k, i in COLS.items()}
        for k in ("isp", "classroom_broad_pct", "nonclassroom_pct",
                  "administration_pct", "isp_current", "isp_prior_year",
                  "isp_highest", "isp_lowest"):
            if isinstance(d.get(k), float):
                d[k] = round(d[k], 4)
        return d

    state = pack(rows[2])
    districts = []
    for row in rows[3:]:
        if not row or row[COLS["county"]] != "Pima":
            continue
        d = pack(row)
        name = d["district"]
        d["short_name"] = COVERED.get(name)
        d["covered_beat"] = name in COVERED
        # derived here, never asked of a model
        isp, st_isp = _num(d["isp"]), _num(state["isp"])
        d["isp_vs_state_pts"] = round((isp - st_isp) * 100, 1) if isp and st_isp else None
        ps, st_ps = _num(d["per_student_total"]), _num(state["per_student_total"])
        d["per_student_vs_state"] = round(ps - st_ps) if ps and st_ps else None
        sal, st_sal = _num(d["teacher_salary"]), _num(state["teacher_salary"])
        d["teacher_salary_vs_state_calc"] = round(sal - st_sal) if sal and st_sal else None
        districts.append(d)

    districts.sort(key=lambda x: -(x["students"] or 0))
    return state, districts


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--no-fetch", action="store_true", help="reuse cached workbook")
    ap.add_argument("--summary", action="store_true")
    a = ap.parse_args()

    path = fetch(force=not a.no_fetch)
    state, districts = extract(path)

    payload = {
        "fiscal_year": FISCAL_YEAR,
        "retrieved": date.today().isoformat(),
        "source": {"name": "Arizona Auditor General, School District Spending",
                   "statute": "A.R.S. § 41-1279.03",
                   "landing": LANDING, "file": XLSX_URL},
        "metric_note": ("'isp' is the Instructional Spending Percentage — the "
                        "Auditor General's headline 'cents of every dollar in the "
                        "classroom'. 'classroom_broad_pct' is a wider definition "
                        "(instruction + student support + instruction support) and "
                        "is NOT the number other outlets quote."),
        "state": state,
        "districts": districts,
    }
    DATA_DIR.mkdir(exist_ok=True)
    OUT_JSON.write_text(json.dumps(payload, indent=1, default=str))
    print(f"wrote {OUT_JSON}  ({len(districts)} Pima districts, FY{FISCAL_YEAR})")

    if a.summary:
        s = state

        def money(v):
            return f"${v:,.0f}" if isinstance(v, (int, float)) else str(v)

        print(f"\nSTATEWIDE  ISP {s['isp']:.1%}  (FY{s['isp_lowest_fy']} is the lowest "
              f"since 2001; peak {s['isp_highest']:.1%} in FY{s['isp_highest_fy']})")
        print(f"           {money(s['per_student_total'])}/student   "
              f"avg teacher salary {money(s['teacher_salary'])}\n")
        print(f"{'district':<34}{'students':>9}{'ISP':>8}{'vs st':>7}"
              f"{'$/student':>11}{'salary':>9}{'s:t':>6}")
        print("-" * 84)
        def fmt(v, spec, width):
            try:
                return format(v, spec)
            except (TypeError, ValueError):
                return format("n/a", f">{width}")

        for d in districts:
            if not d["covered_beat"]:
                continue
            print(f"{(d['short_name'] or d['district'])[:33]:<34}"
                  f"{fmt(d['students'], '>9,', 9)}"
                  f"{fmt(d['isp'], '>8.1%', 8)}"
                  f"{fmt(d['isp_vs_state_pts'], '>+7.1f', 7)}"
                  f"{fmt(d['per_student_total'], '>11,', 11)}"
                  f"{fmt(d['teacher_salary'], '>9,', 9)}"
                  f"{fmt(d['students_per_teacher'], '>6.1f', 6)}")


if __name__ == "__main__":
    main()
